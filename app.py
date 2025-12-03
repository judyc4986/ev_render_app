from flask import Flask, render_template, request
import openpyxl
import os

app = Flask(__name__)

# =========================================================
# Load statewide supercharger count WITHOUT pandas
# =========================================================
def load_statewide_supercharger_count():
    excel_path = "supercharger_by_county_summary.xlsx"  # Must be in project root

    if not os.path.exists(excel_path):
        print("⚠ Missing supercharger_by_county_summary.xlsx")
        return None

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb.active

        # Find the header columns
        county_col = None
        sc_col = None

        for col in range(1, ws.max_column + 1):
            header = str(ws.cell(row=1, column=col).value).strip().lower()
            if header == "county":
                county_col = col
            if "super" in header or "charger" in header:
                sc_col = col

        if not county_col or not sc_col:
            print("⚠ Required columns not found")
            return None

        # Find the row where county == "total"
        for row in range(2, ws.max_row + 1):
            county_val = str(ws.cell(row=row, column=county_col).value).strip().lower()
            if county_val == "total":
                sc_value = ws.cell(row=row, column=sc_col).value
                return int(sc_value)

        print("⚠ Row 'Total' not found")
        return None

    except Exception as e:
        print("⚠ Error loading statewide count:", e)
        return None


STATEWIDE_SC = load_statewide_supercharger_count()


# =========================================================
# Statewide forecast model
# =========================================================
def statewide_forecast(x):
    ev_reg = (
        0.0026120043 * (x ** 3)
        - 4.7343743373 * (x ** 2)
        + 4547.3532407731 * x
        - 140511.0057756579
    )

    adopt = (
        0.0000000007 * (x ** 3)
        - 0.0000012533 * (x ** 2)
        + 0.0012038019 * x
        - 0.0371968938
    )

    return ev_reg, adopt


@app.route("/", methods=["GET", "POST"])
def index():
    result = None
    error = None

    if request.method == "POST":
        try:
            x_raw = request.form.get("superchargers", "").strip()
            x = float(x_raw)

            ev_reg, adopt = statewide_forecast(x)

            result = {
                "x": x,
                "ev_reg_str": f"{ev_reg:,.0f}",
                "adopt_pct_str": f"{adopt * 100:.2f}",
            }
        except ValueError:
            error = "Please enter a valid numeric value."

    return render_template(
        "index.html",
        result=result,
        error=error,
        statewide_sc=STATEWIDE_SC,
    )


if __name__ == "__main__":
    app.run(debug=True)
